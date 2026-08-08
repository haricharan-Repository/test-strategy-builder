import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@dataclass
class TestCase:
    category: str
    title: str
    description: str
    inputs: str
    expected_result: str
    priority: str
    notes: str


class TestStrategyBuilder:
    def __init__(self, requirement: str):
        self.requirement = requirement.strip()
        self.sentences = self._split_sentences(self.requirement)
        self.keywords = self._extract_keywords(self.requirement)

    def _split_sentences(self, text: str) -> List[str]:
        parts = re.split(r'(?<=[.!?])\s+', text.strip())
        return [p.strip() for p in parts if p.strip()]

    def _extract_keywords(self, text: str) -> List[str]:
        words = re.findall(r"\b[a-zA-Z]{4,}\b", text.lower())
        stopwords = {
            'system', 'application', 'user', 'users', 'should', 'must', 'will', 'also', 'when', 'then', 'with',
            'that', 'from', 'their', 'them', 'there', 'this', 'these', 'those', 'any', 'each', 'many', 'more',
            'less', 'have', 'has', 'been', 'being', 'are', 'is', 'be', 'can', 'could', 'would', 'should', 'must'
        }
        filtered = [w for w in words if w not in stopwords]
        freq = {}
        for word in filtered:
            freq[word] = freq.get(word, 0) + 1
        sorted_keywords = sorted(freq.keys(), key=lambda k: (-freq[k], k))
        return sorted_keywords[:12]

    def build(self) -> List[TestCase]:
        cases = []
        cases.extend(self._build_core_scenarios())
        cases.extend(self._build_edge_cases())
        cases.extend(self._build_negative_tests())
        cases.extend(self._build_additional_tests())
        return cases

    def _build_core_scenarios(self) -> List[TestCase]:
        title = 'Core functional path'
        description = 'Validate the primary business flow described by the requirement.'
        inputs = self.requirement
        expected = 'The system behaves according to the requirement without errors.'
        return [TestCase('Core Scenario', title, description, inputs, expected, 'High', 'Main happy path')]

    def _build_edge_cases(self) -> List[TestCase]:
        results = []
        # Generate edge case testing hints based on value words
        results.append(TestCase(
            'Edge Case',
            'Boundary values and limits',
            'Check behavior at the minimum, maximum, and just outside valid ranges for the important inputs.',
            self._keyword_edge_inputs(),
            'The system accepts valid boundary values and rejects values outside the expected range with clear feedback.',
            'High',
            'Includes limit conditions and extreme but valid values'
        ))
        results.append(TestCase(
            'Edge Case',
            'Missing optional data or blank fields',
            'Confirm the system handles omitted or blank optional data without failing.',
            'Blank values for optional fields and omitted optional sections.',
            'The system continues smoothly and uses defaults or prompts as required.',
            'Medium',
            'Helps verify graceful degradation when data is absent'
        ))
        return results

    def _keyword_edge_inputs(self) -> str:
        if not self.keywords:
            return 'Minimum and maximum lengths, values, dates, or amounts based on the requirement.'
        return ' / '.join([f'{kw} at minimum, maximum, and just outside valid range' for kw in self.keywords[:4]])

    def _build_negative_tests(self) -> List[TestCase]:
        cases = []
        cases.append(TestCase(
            'Negative Test',
            'Invalid input formats',
            'Submit inputs using incorrect formats, invalid characters, or unsupported values.',
            'Malformed data, invalid number formats, wrong date formats, and unsupported characters.',
            'The system rejects invalid input with meaningful error messages and prevents incorrect processing.',
            'High',
            'Focus on validation and error handling for user input'
        ))
        cases.append(TestCase(
            'Negative Test',
            'Unauthorized access or invalid user role',
            'Attempt to perform the action with a user who lacks the required permission or role.',
            'User account with insufficient privileges, access denied conditions.',
            'Access is denied, and the system logs the attempt without exposing sensitive data.',
            'High',
            'Security-focused negative scenario'
        ))
        cases.append(TestCase(
            'Negative Test',
            'Failure in external dependency',
            'Simulate a problem with external services or integrations required for the flow.',
            'Timeouts, service errors, unavailable APIs.',
            'The system handles the failure gracefully and displays an appropriate error or retry path.',
            'Medium',
            'Validates resiliency and failure recovery'
        ))
        return cases

    def _build_additional_tests(self) -> List[TestCase]:
        cases = []
        cases.append(TestCase(
            'Usability Test',
            'Ease of use and user guidance',
            'Review the flow from the end-user perspective to ensure the steps are intuitive and instructions are clear.',
            'User navigation through the feature with screen labels, buttons, and prompts.',
            'Users can complete the task without confusion and can recover from mistakes.',
            'Medium',
            'Important for user-facing functionality'
        ))
        cases.append(TestCase(
            'Performance Test',
            'Load and response time',
            'Measure performance when the feature is used under normal and heavy load.',
            'Normal usage plus surge conditions or high transaction volume.',
            'Performance stays within acceptable limits and does not degrade critically under stress.',
            'Medium',
            'Helps identify bottlenecks early'
        ))
        cases.append(TestCase(
            'Regression Test',
            'Previous functionality remains intact',
            'Verify that related existing features continue to work after implementing the new requirement.',
            'Repeat key existing workflows connected to this requirement.',
            'No regressions occur in related areas.',
            'Medium',
            'Covers the integration of new behavior with existing systems'
        ))
        return cases

    def to_excel(self, output_path: Path) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = 'Requirement Summary'
        summary_sheet.append(['Requirement Text'])
        summary_sheet.append([self.requirement])
        summary_sheet.append([])
        summary_sheet.append(['Extracted Keywords'])
        summary_sheet.append([', '.join(self.keywords) or 'N/A'])

        strategy_sheet = workbook.create_sheet('Test Strategy')
        headers = ['Category', 'Test Case', 'Description', 'Inputs', 'Expected Result', 'Priority', 'Notes']
        strategy_sheet.append(headers)
        for case in self.build():
            strategy_sheet.append([
                case.category,
                case.title,
                case.description,
                case.inputs,
                case.expected_result,
                case.priority,
                case.notes,
            ])

        self._auto_fit_columns(summary_sheet)
        self._auto_fit_columns(strategy_sheet)
        workbook.save(output_path)

    def _auto_fit_columns(self, sheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            sheet.column_dimensions[column].width = adjusted_width


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Generate an Excel-based test strategy from a requirement description.'
    )
    parser.add_argument('-r', '--requirement-file', type=Path, help='Path to a text file containing the requirement.')
    parser.add_argument('-o', '--output', type=Path, default=Path('test_strategy.xlsx'), help='Output Excel filename.')
    parser.add_argument('-t', '--title', type=str, default='Test Strategy', help='Optional title for the Excel output (stored in the workbook).')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.requirement_file and args.requirement_file.exists():
        requirement = args.requirement_file.read_text(encoding='utf-8').strip()
    else:
        requirement = input('Enter the requirement text: ').strip()

    if not requirement:
        raise SystemExit('Requirement text cannot be empty.')

    builder = TestStrategyBuilder(requirement)
    builder.to_excel(args.output)
    print(f'Created test strategy workbook: {args.output}')


if __name__ == '__main__':
    main()
